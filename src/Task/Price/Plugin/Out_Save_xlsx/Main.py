# Created: 2022.10.14
# Author: Vladimir Vons <VladVons@gmail.com>
# License: GNU, see LICENSE for more details


import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#
from Inc.DbList import TDbList
from Inc.Util.Obj import GetClassPath
from Inc.ParserX.Common import TFileBase
from IncP.Log import Log


class TMain(TFileBase):
    def DblCreateXlsx(self, aParam: list):
        WB = Workbook()
        WB.remove_sheet(WB.active)

        for Param in aParam:
            Dbl: TDbList = Param['dbl']
            WS = WB.create_sheet(title = Param['sheet'])

            Fields = Param.get('fields', Dbl.GetFields())
            RowNo = 1
            Rec = Dbl.RecGo(0)
            for Idx, Field in enumerate(Fields):
                CD = WS.column_dimensions[get_column_letter(Idx + 1)]

                Cell = WS.cell(RowNo, Idx + 1)
                Cell.value = Field
                Cell.font = Font(bold = True)
                if (isinstance(Rec.GetField(Field), (int, float))):
                    CD.number_format = '#,##0.00'

            RowNo += 1
            WS.freeze_panes = WS.cell(RowNo, 1)
            for Rec in Dbl:
                for Idx, Field in enumerate(Fields):
                    Cell = WS.cell(RowNo + Dbl.RecNo, Idx + 1).value = Rec.GetField(Field)
        return WB

    async def Save(self, aParam: dict):
        ConfFile = self.Parent.GetFile()
        Dir = ConfFile.rsplit('.', maxsplit=1)[0]
        os.makedirs(Dir, exist_ok=True)

        for PluginKey, PluginVal in aParam.items():
            ParamExport = self.Parent.GetParamExport(PluginKey)
            Dbls = []
            for ParamKey, ParamVal in PluginVal.items():
                ClassPath = GetClassPath(ParamVal)
                if ('/TDbList' in ClassPath):
                    Param = ParamExport.get(ParamKey, {})
                    Param['sheet'] = ParamKey
                    Param['dbl'] = ParamVal
                    Dbls.append(Param)
                else:
                    Log.Print(1, 'i', f'Unknown class {ClassPath}')
                await self.Sleep.Update()

            if (Dbls):
                File = f'{Dir}/{PluginKey}.xlsx'
                WB = self.DblCreateXlsx(Dbls)
                WB.save(File)

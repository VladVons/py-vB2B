# Created: 2022.10.13
# Author: Vladimir Vons <VladVons@gmail.com>
# License: GNU, see LICENSE for more details


import os
#--- xlsx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.comments import Comment
#
from Inc.ParserX.Common import TFileBase
from IncP.Log import Log
from ..CommonDb import TDbPriceJoin, TDbPrice


class TMain(TFileBase):
    def Head(self, aDbl: TDbPrice, aWS, aHeadRow: dict):
        ConfFormat = self.Parent.Conf.GetKey('format', '#,##0.00')
        ConfFieldWidth = self.Parent.Conf.GetKey('field.width', {})

        for Key, (No, Field, _) in aDbl.Fields.items():
            aWS.cell(aHeadRow['title'], No + 1).value = Key
            CD = aWS.column_dimensions[get_column_letter(No + 1)]

            Width = ConfFieldWidth.get(Key)
            if (Width is not None):
                if (Width == 0):
                    CD.hidden= True
                else:
                    CD.width = Width

            if (Key.startswith('price')):
                CD.number_format = ConfFormat

        Field = aDbl.Fields.get('price')
        #aWS.column_dimensions[get_column_letter(Field[0] + 1)].font = Font(bold=True)
        aWS.freeze_panes = aWS.cell(aHeadRow['data'], Field[0] + 2)

    async def Save(self, aDbPriceJoin: TDbPriceJoin, aPrices: list[TDbPrice]):
        WB = Workbook()
        WS = WB.active

        HeadRow = {'title': 1, 'data': 2}
        self.Head(aDbPriceJoin, WS, HeadRow)

        ConfRatio = self.Parent.Conf.GetKey('ratio')
        for RowNo, Rec in enumerate(aDbPriceJoin, HeadRow['data']):
            Price = Rec.price
            for FieldNo, Field in enumerate(aDbPriceJoin.Fields):
                Value = Rec[FieldNo]
                Cell = WS.cell(RowNo, FieldNo + 1)
                Cell.value = Value
                if (Field.startswith('In_Price_')):
                    if (Value == Price):
                        Cell.font = Font(bold = True)
                    elif (ConfRatio):
                        Text = '+%0.2f %0.2f%%' % (Value - Price, Value / Price * 100 - 100)
                        Cell.comment = Comment(Text, '', 20)
            await self.Sleep.Update()

        ConfPrices = self.Parent.Conf.GetKey('prices', True)
        if (ConfPrices):
            aDbPriceJoin.SearchAdd('mpn')
            for Plugin, Dbl in aPrices.items():
                WS = WB.create_sheet(title = Plugin)
                self.Head(Dbl, WS, HeadRow)
                for RowNo, Rec in enumerate(Dbl, HeadRow['data']):
                    for FieldNo, Field in enumerate(Dbl.Fields):
                        Cell = WS.cell(RowNo, FieldNo + 1)
                        Cell.value = Rec[FieldNo]
                        if (Field == 'price') and (aDbPriceJoin.Search('mpn', Rec.mpn) >= 0):
                            Cell.font = Font(bold = True)
                    await self.Sleep.Update()

        ConfFile = self.Parent.GetFile()
        os.makedirs(os.path.dirname(ConfFile), exist_ok=True)
        Log.Print(1, 'i', 'Save %s'  % (ConfFile))
        WB.save(ConfFile)
